from openpyxl import *
from openpyxl.styles import *
from openpyxl.chart import *

# Global variables
kwBook = load_workbook(filename = "kw_report.xlsx")
kwSheet = kwBook.active 

def removeColumnPrefixes():
    numCol = 10    

    # For each column in the first row, replace the "ns1:" prefix
    for col in range(1, numCol + 1):
        c = kwSheet.cell(1, col)
        c.value = str(c.value).replace("ns1:", "")

def removeCellColors():
    white = "00FFFFFF"

    # For all cells, set the fill type to be none
    for row in kwSheet:
        for cell in row: 
            cell.fill = PatternFill(start_color = white, end_color = white, fill_type = None)

def ModifyFilePath():
    pathColumn = 2

    # For each second cell after row two, delete all characters before "pmic_drv" 
    for row in range(2, kwSheet.max_row + 1):
        c = kwSheet.cell(row, pathColumn)
        
        index = str(c.value).find("pmic_drv")
        if (index != -1):
            c.value = str(c.value)[index:]
            
def modifyColumns():
    # Delete columns
    kwSheet.delete_cols(7)   # citingStatus
    kwSheet.delete_cols(9)   # displayAs
    kwSheet.delete_cols(10)  # metaInf

    # Add disposition and comment columns
    kwSheet.cell(1, kwSheet.max_column + 1).value = "disposition"
    kwSheet.cell(1, kwSheet.max_column + 1).value = "comment"

def createKwSummary():
    # Constant; number of actual Klocwork issues; varies with every Klocwork report
    numIssues = kwSheet.max_row
    # Constants; Klocwork summary report columns; stays the same regardless of Klocwork report
    kwSumErrCodeCol = 1 
    kwSumOccurCol = 2 
    kwSumSevCol = 3 
    kwSumSevLevCol = 4
    kwSumCheckCol = 5 
    kwSumDisCol = 6 
    kwSumComCol = 7 
    # Constants; columns of an actual Klocwork issue; stays the same regardless of Klocwork report
    errCodeCol = 5
    sevCol = 7
    sevLevCol = 8
    checkCol = 9

    # Create title of the summary table 
    kwSheet.merge_cells(start_row = kwSheet.max_row + 3, start_column = kwSumErrCodeCol, end_row = kwSheet.max_row + 3, end_column = kwSumComCol)
    kwSheet.cell(kwSheet.max_row, 1).alignment = Alignment(horizontal = "center")
    kwSheet.cell(kwSheet.max_row, 1).value = "klocwork issue summary"
    
    # Create stat report fields 
    kwSheet.cell(kwSheet.max_row + 1, kwSumErrCodeCol).value = "error code"
    kwSheet.cell(kwSheet.max_row, kwSumOccurCol).value = "occurrences"
    kwSheet.cell(kwSheet.max_row, kwSumSevCol).value = "severity"
    kwSheet.cell(kwSheet.max_row, kwSumSevLevCol).value = "severity level"
    kwSheet.cell(kwSheet.max_row, kwSumCheckCol).value = "checker"
    kwSheet.cell(kwSheet.max_row, kwSumDisCol).value = "disposition"
    kwSheet.cell(kwSheet.max_row, kwSumComCol).value = "comment"
    kwSumFirstEntryRow = kwSheet.max_row + 1

    # For each Klocwork issue, get its error code and compare it to existing error codes in the summary
    # report table. If there was no match, make an entry in the summary table for the Klocwork issue.
    # Else if there is a match, increment the number of occurrence
    for issue in range(2, numIssues + 1):
        errCode = kwSheet.cell(issue, errCodeCol).value
        
        kwSumEntryRow = kwSumFirstEntryRow
        while (1):
            kwSumErrCode = kwSheet.cell(kwSumEntryRow, kwSumErrCodeCol).value
            if (kwSumErrCode is None):
                kwSheet.cell(kwSumEntryRow, kwSumErrCodeCol).value = errCode
                kwSheet.cell(kwSumEntryRow, kwSumOccurCol).value = 1
                kwSheet.cell(kwSumEntryRow, kwSumSevCol).value = kwSheet.cell(issue, sevCol).value
                kwSheet.cell(kwSumEntryRow, kwSumSevLevCol).value = kwSheet.cell(issue, sevLevCol).value
                kwSheet.cell(kwSumEntryRow, kwSumCheckCol).value = kwSheet.cell(issue, checkCol).value
                break
            elif (str(errCode) == str(kwSumErrCode)):
                kwSheet.cell(kwSumEntryRow, kwSumOccurCol).value += 1
                break
            kwSumEntryRow += 1
    return kwSumFirstEntryRow

def createKwSummaryChart(kwSumFirstEntryRow):
    data = Reference(kwSheet, min_col = 2, min_row = kwSumFirstEntryRow - 1, max_col = 2, max_row = kwSheet.max_row)
    titles = Reference(kwSheet, min_col = 1, min_row = kwSumFirstEntryRow - 1, max_col = 1, max_row = kwSheet.max_row)

    chart = BarChart3D()
    chart.title = "Klocwork Issues"
    chart.add_data(data = data, titles_from_data = True)
    chart.set_categories(titles)
    chart.height = 20
    chart.width = 40

    kwSheet.add_chart(chart, "A" + str(kwSheet.max_row + 1))
            
# Execution
if __name__ == "__main__":
    # Remove 'ns1:' prefix in all columns within the first row
    removeColumnPrefixes()

    # Remove colors from all cells
    removeCellColors()

    # Modify file paths to start at pmic_drv
    ModifyFilePath()

    # Delete citingStatus, displayAs, and metaInf columns; add columns for dispositions and comments 
    modifyColumns()

    # Display frequency of errors along with their severity, severity level, and checker
    kwSumFirstEntryRow = createKwSummary()

    # Display graphical representation of the Klocwork summary
    createKwSummaryChart(kwSumFirstEntryRow)

    kwBook.save("burton_kw_report.xlsx")